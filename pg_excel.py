import psycopg2
import xlrd
import sys
from openpyxl import load_workbook

#function create the contacts table 
def create_table(cur):
    cur.execute("""
        CREATE TABLE IF NOT EXISTS contacts (
            Name VARCHAR,
            Country VARCHAR,
            Email VARCHAR,
            Phone VARCHAR,
            Active BOOLEAN
        )
    """)

    
#function read and insert data into POSTGRESQL 
def read_excel_and_insert_data(cur, filename):
    wb = load_workbook(filename)
    sheet = wb.active
    
    #iterate through rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, country, email, phone, active = row
        phone = str(phone) if phone else ''  #handle None values
        active = bool(active) if active is not None else False  #handle None values
        #executing sql insert query
        cur.execute("""
            INSERT INTO contacts (Name, Country, Email, Phone, Active)
            VALUES (%s, %s, %s, %s, %s)
        """, (name, country, email, phone, active))


def main():
    #check if correct number of command 
    if len(sys.argv) != 4:
        print("Usage: python pg_excel.py <username> <password> <database_name>")
        sys.exit(1)
    # extract username, password and database name from command
    username, password, dbname = sys.argv[1], sys.argv[2], sys.argv[3]

    try:
        # Connect to POSTGRESQL
        conn = psycopg2.connect(
            dbname=dbname,
            user=username,
            password=password
        )
        cur = conn.cursor()

        create_table(cur) #create table if not exist
        read_excel_and_insert_data(cur, "Contacts.xlsx") #read and insert data from Contacts.xlsx

        conn.commit() #commit the transaction
        print("Data inserted successfully!")

    except psycopg2.Error as e:
        print("Error connecting to PostgreSQL:", e)
        sys.exit(1)

    finally:
        if conn:
            conn.close()

if __name__ == "__main__":
    main()
